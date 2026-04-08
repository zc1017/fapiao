import sys
import os
import re
import io
import zipfile
import base64
import tempfile
import urllib.request
import urllib.error
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QLabel, QListWidget, QListWidgetItem,
    QSplitter, QProgressBar, QStatusBar, QDialog, QCheckBox, QScrollArea, QTextEdit, QStyledItemDelegate, QStyleOptionViewItem
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QRect, QSize
from PyQt5.QtGui import QFont, QColor, QIcon, QPixmap, QPainter, QPen, QBrush
import cv2
import numpy as np
from pyzbar.pyzbar import decode, ZBarSymbol
from defusedxml import ElementTree as ET
import openpyxl
from openpyxl.styles import Font as ExcelFont, Alignment, Border, Side
import fitz
from PIL import Image
import io

_ocr_instance = None
_log_messages = []
_file_logs = {}
_log_lock = threading.Lock()
_ocr_lock = threading.Lock()

def add_log(message, file_path=None):
    """添加日志消息（线程安全）"""
    global _log_messages, _file_logs, _log_lock
    with _log_lock:
        _log_messages.append(message)
        if file_path:
            if file_path not in _file_logs:
                _file_logs[file_path] = []
            _file_logs[file_path].append(message)
    print(message)

def get_logs():
    """获取所有日志消息（线程安全）"""
    global _log_messages, _log_lock
    with _log_lock:
        return '\n'.join(_log_messages)

def get_file_logs(file_path):
    """获取指定文件的日志消息（线程安全）"""
    global _file_logs, _log_lock
    with _log_lock:
        return '\n'.join(_file_logs.get(file_path, []))

def clear_logs():
    """清空日志消息（线程安全）"""
    global _log_messages, _file_logs, _log_lock
    with _log_lock:
        _log_messages = []
        _file_logs = {}

def get_ocr():
    """获取OCR实例（线程安全）"""
    global _ocr_instance, _ocr_lock
    with _ocr_lock:
        if _ocr_instance is None:
            try:
                from rapidocr_onnxruntime import RapidOCR
                _ocr_instance = RapidOCR()
                add_log("RapidOCR初始化成功")
            except Exception as e:
                add_log(f"RapidOCR初始化失败: {e}")
                return None
        return _ocr_instance


class StatusTagDelegate(QStyledItemDelegate):
    """状态列的Tag样式委托"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
    
    def paint(self, painter, option, index):
        """绘制Tag样式的状态"""
        painter.save()
        
        text = index.data(Qt.DisplayRole)
        if not text:
            painter.restore()
            super().paint(painter, option, index)
            return
        
        if text == '成功':
            bg_color = QColor(76, 175, 80)
            text_color = QColor(255, 255, 255)
            border_color = QColor(200, 230, 201)
        else:
            bg_color = QColor(245, 108, 108)
            text_color = QColor(255, 255, 255)
            border_color = QColor(255, 179, 172)
        
        rect = option.rect
        padding = 8
        tag_height = 24
        tag_width = min(60, rect.width() - padding * 2)
        
        tag_rect = QRect(
            rect.x() + (rect.width() - tag_width) // 2,
            rect.y() + (rect.height() - tag_height) // 2,
            tag_width,
            tag_height
        )
        
        painter.setRenderHint(QPainter.Antialiasing)
        
        painter.setBrush(QBrush(bg_color))
        painter.setPen(QPen(border_color, 1))
        painter.drawRoundedRect(tag_rect, 4, 4)
        
        painter.setPen(text_color)
        font = painter.font()
        font.setPointSize(9)
        font.setBold(True)
        painter.setFont(font)
        
        painter.drawText(tag_rect, Qt.AlignCenter, text)
        
        painter.restore()
    
    def sizeHint(self, option, index):
        """返回合适的大小"""
        return QSize(70, 30)


class InvoiceParser:
    """发票解析器"""
    
    @staticmethod
    def fetch_invoice_xml_from_url(url):
        """从发票二维码URL获取XML数据"""
        try:
            if not url.startswith('http'):
                return None, '无效的URL'
            
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Accept': '*/*',
                    'Accept-Language': 'zh-CN,zh;q=0.9',
                }
            )
            
            with urllib.request.urlopen(req, timeout=10) as response:
                content = response.read()
                
                if content[:4] == b'PK\x03\x04':
                    with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                        xml_files = [f for f in zf.namelist() if f.endswith('.xml')]
                        if xml_files:
                            xml_content = zf.read(xml_files[0])
                            try:
                                return xml_content.decode('utf-8'), None
                            except:
                                return xml_content.decode('gbk'), None
                
                try:
                    xml_content = content.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        xml_content = content.decode('gbk')
                    except:
                        xml_content = content.decode('utf-8', errors='ignore')
                
                if '<?xml' in xml_content.lower() or '<EInvoice' in xml_content:
                    return xml_content, None
                
                if re.match(r'^[A-Za-z0-9+/=]+$', xml_content.strip()) and len(xml_content.strip()) > 50:
                    decoded = InvoiceParser.decode_base64_to_xml(xml_content.strip())
                    if decoded:
                        return decoded, None
                
                return None, f'URL返回的数据格式无法识别: {xml_content[:100]}...'
                
        except urllib.error.HTTPError as e:
            return None, f'HTTP错误: {e.code}'
        except urllib.error.URLError as e:
            return None, f'网络错误: {str(e.reason)}'
        except Exception as e:
            return None, f'获取数据失败: {str(e)}'
    
    @staticmethod
    def ocr_extract_seller_info(image_path, existing_total_price=None):
        """使用OCR提取销售方和购买方信息"""
        try:
            ocr = get_ocr()
            if ocr is None:
                return None, 'OCR未初始化'
            
            if image_path.lower().endswith('.pdf'):
                doc = fitz.open(image_path)
                page = doc.load_page(0)
                
                text_content = page.get_text()
                if text_content.strip():
                    add_log(f"PDF文本层内容:\n{text_content}\n", image_path)
                
                mat = fitz.Matrix(3.0, 3.0)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                nparr = np.frombuffer(img_data, np.uint8)
                image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                doc.close()
            else:
                image = cv2.imread(image_path)
            
            if image is None:
                return None, '无法读取图片'
            
            add_log(f"\n开始OCR识别: {os.path.basename(image_path)}", image_path)
            
            result, elapse = ocr(image)
            if isinstance(elapse, list):
                total_elapse = sum(elapse) if elapse else 0
                add_log(f"OCR识别耗时: {total_elapse:.2f}ms", image_path)
            else:
                add_log(f"OCR识别耗时: {elapse:.2f}ms", image_path)
            
            if not result:
                add_log("OCR未识别到任何结果", image_path)
                return None, 'OCR未识别到文字'
            
            add_log(f"识别到 {len(result)} 条文字", image_path)
            
            all_text = []
            for item in result:
                if item and isinstance(item, (list, tuple)) and len(item) >= 2:
                    box = item[0]
                    text_info = item[1]
                    if isinstance(text_info, (list, tuple)) and len(text_info) >= 1:
                        text = str(text_info[0]).strip()
                        confidence = float(text_info[1]) if len(text_info) > 1 else 1.0
                    else:
                        text = str(text_info).strip()
                        confidence = 1.0
                    if text:
                        all_text.append(text)
                        add_log(f"  {text} (置信度: {confidence:.2f})", image_path)
            
            full_text = '\n'.join(all_text)
            add_log(f"\nOCR识别到的所有文字:\n{full_text}\n", image_path)
            
            info = {}
            
            names = []
            tax_ids = []
            
            full_width_map = str.maketrans('０１２３４５６７８９', '0123456789')
            
            for text in all_text:
                if '名称' in text:
                    name_match = re.search(r'名称[：:\s]*(.+)', text)
                    if name_match:
                        name = name_match.group(1).strip()
                        if name and len(name) > 2:
                            names.append(name)
                
                if '统一社会信用代码' in text or '纳税人识别号' in text:
                    text_normalized = text.translate(full_width_map)
                    tax_match = re.search(r'[：:\s]*([0-9A-Z]{15,20})', text_normalized)
                    if tax_match:
                        tax_id = tax_match.group(1).strip()
                        if tax_id:
                            tax_ids.append(tax_id)
            
            all_tax_pattern = r'[0-9A-Z]{15,20}'
            all_tax_ids = re.findall(all_tax_pattern, full_text.translate(full_width_map))
            
            for tax_id in all_tax_ids:
                if tax_id not in tax_ids:
                    tax_ids.append(tax_id)
            
            add_log(f"识别到的名称: {names}", image_path)
            add_log(f"识别到的税号: {tax_ids}", image_path)
            
            if len(names) >= 1:
                info['购买方名称'] = names[0]
            if len(names) >= 2:
                info['销售方名称'] = names[1]
            
            if len(tax_ids) >= 1:
                info['购买方纳税人识别号'] = tax_ids[0]
            if len(tax_ids) >= 2:
                info['销售方纳税人识别号'] = tax_ids[1]
            
            amounts = []
            for text in all_text:
                if '￥' in text or '¥' in text:
                    patterns = [
                        r'[￥¥]\s*([\d,\s]+\.?\s*\d*)',
                        r'([\d,\s]+\.?\s*\d*)\s*[￥¥]',
                    ]
                    for pattern in patterns:
                        matches = re.findall(pattern, text)
                        for match in matches:
                            try:
                                amount_str = match.replace(',', '').replace(' ', '').strip()
                                if amount_str and '.' in amount_str:
                                    amount = float(amount_str)
                                    if amount > 0:
                                        amounts.append(amount)
                                        add_log(f"  识别到金额: {amount}", image_path)
                            except:
                                pass
            
            amounts = list(set(amounts))
            amounts.sort(reverse=True)
            add_log(f"所有识别到的金额: {amounts}", image_path)
            
            found = False
            
            if existing_total_price and len(amounts) >= 2:
                total_price_value = float(existing_total_price)
                add_log(f"已有价税合计: {total_price_value}, 尝试匹配合计金额和合计税额", image_path)
                
                for i in range(len(amounts)):
                    if abs(amounts[i] - total_price_value) < 0.01:
                        for j in range(len(amounts)):
                            if i != j:
                                total_amount = amounts[j]
                                total_tax = total_price_value - total_amount
                                if total_tax >= 0:
                                    info['合计金额'] = str(total_amount)
                                    info['合计税额'] = str(round(total_tax, 2))
                                    info['价税合计'] = str(total_price_value)
                                    add_log(f"匹配成功（价税合计匹配）: 合计金额={total_amount}, 合计税额={round(total_tax, 2)}, 价税合计={total_price_value}", image_path)
                                    found = True
                                    break
                        if found:
                            break
                
                if not found:
                    for i in range(len(amounts)):
                        for j in range(i + 1, len(amounts)):
                            total_amount = amounts[i]
                            total_tax = amounts[j]
                            
                            if abs(total_amount + total_tax - total_price_value) < 0.01:
                                if total_amount > total_tax:
                                    info['合计金额'] = str(total_amount)
                                    info['合计税额'] = str(total_tax)
                                    info['价税合计'] = str(total_price_value)
                                    add_log(f"匹配成功: 合计金额={total_amount}, 合计税额={total_tax}, 价税合计={total_price_value}", image_path)
                                    found = True
                                    break
                        if found:
                            break
                
                if not found:
                    for i in range(len(amounts)):
                        for j in range(i + 1, len(amounts)):
                            total_tax = amounts[i]
                            total_amount = amounts[j]
                            
                            if abs(total_amount + total_tax - total_price_value) < 0.01:
                                info['合计金额'] = str(total_amount)
                                info['合计税额'] = str(total_tax)
                                info['价税合计'] = str(total_price_value)
                                add_log(f"匹配成功: 合计金额={total_amount}, 合计税额={total_tax}, 价税合计={total_price_value}", image_path)
                                found = True
                                break
                        if found:
                            break
            
            if not found and existing_total_price and len(amounts) >= 1:
                total_price_value = float(existing_total_price)
                for amount in amounts:
                    if abs(amount - total_price_value) < 0.01:
                        info['合计金额'] = str(amount)
                        info['合计税额'] = '0'
                        info['价税合计'] = str(total_price_value)
                        add_log(f"匹配成功（税额为0）: 合计金额={amount}, 合计税额=0, 价税合计={total_price_value}", image_path)
                        found = True
                        break
            
            if not found and len(amounts) >= 3:
                for i in range(len(amounts)):
                    for j in range(i + 1, len(amounts)):
                        for k in range(j + 1, len(amounts)):
                            total_amount = amounts[i]
                            total_tax = amounts[j]
                            total_price = amounts[k]
                            
                            if abs(total_amount + total_tax - total_price) < 0.01:
                                if total_amount > total_tax:
                                    info['合计金额'] = str(total_amount)
                                    info['合计税额'] = str(total_tax)
                                    info['价税合计'] = str(total_price)
                                    add_log(f"匹配成功: 合计金额={total_amount}, 合计税额={total_tax}, 价税合计={total_price}", image_path)
                                    found = True
                                    break
                        if found:
                            break
                    if found:
                        break
                
                if not found:
                    add_log("未找到满足条件的金额组合", image_path)
            
            add_log(f"\n提取的信息: {info}", image_path)
            
            return info if info else None, None
            
        except Exception as e:
            add_log(f"OCR识别错误: {str(e)}", image_path)
            import traceback
            traceback.print_exc()
            return None, f'OCR识别错误: {str(e)}'
            
            names = []
            for i, line in enumerate(lines):
                line = line.strip()
                if is_company_name(line):
                    names.append((i, line))
                elif '名称' in line:
                    name_match = re.search(r'名称[：:\s]*(.+?)(?:\s|$|纳税人|地址|电话|统一)', line)
                    if name_match:
                        name = clean_company_name(name_match.group(1))
                        if len(name) > 2 and is_company_name(name):
                            names.append((i, name))
            
            if not names:
                for i, line in enumerate(lines):
                    line = line.strip()
                    if is_company_name(line):
                        names.append((i, line))
            
            company_names = [(idx, name) for idx, name in names if is_company_name(name)]
            
            if len(company_names) >= 2:
                info['销售方名称'] = company_names[-1][1]
                info['购买方名称'] = company_names[0][1]
            elif seller_idx >= 0 and buyer_idx >= 0:
                seller_names = [(idx, name) for idx, name in company_names if seller_idx < idx < buyer_idx]
                buyer_names = [(idx, name) for idx, name in company_names if idx > buyer_idx]
                
                if seller_names:
                    info['销售方名称'] = seller_names[0][1]
                if buyer_names:
                    info['购买方名称'] = buyer_names[0][1]
            elif seller_idx >= 0:
                for idx, name in company_names:
                    if idx > seller_idx and idx < seller_idx + 15:
                        info['销售方名称'] = name
                        break
                
                if '销售方名称' not in info:
                    for idx, name in names:
                        if idx > seller_idx and idx < seller_idx + 15:
                            info['销售方名称'] = name
                            break
            
            if '销售方名称' not in info and company_names:
                info['销售方名称'] = company_names[-1][1]
            
            if tax_ids:
                if seller_idx >= 0 and buyer_idx >= 0:
                    for tax_id in tax_ids:
                        for i, line in enumerate(lines):
                            if tax_id in line:
                                if abs(i - seller_idx) < abs(i - buyer_idx):
                                    info['销售方纳税人识别号'] = tax_id
                                else:
                                    info['购买方纳税人识别号'] = tax_id
                                break
                elif len(tax_ids) >= 2:
                    info['销售方纳税人识别号'] = tax_ids[-1]
                    info['购买方纳税人识别号'] = tax_ids[0]
                else:
                    info['销售方纳税人识别号'] = tax_ids[0]
            
            return info if info else None, None
            
        except Exception as e:
            return None, f'OCR识别错误: {str(e)}'
    
    @staticmethod
    def decode_qrcode(image_path):
        """识别图片或PDF中的二维码"""
        try:
            if image_path.lower().endswith('.pdf'):
                return InvoiceParser.decode_qrcode_from_pdf(image_path)
            
            image = cv2.imread(image_path)
            if image is None:
                return None, '无法读取图片文件'
            
            decoded_objects = decode(image, symbols=[ZBarSymbol.QRCODE])
            
            if not decoded_objects:
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                decoded_objects = decode(gray, symbols=[ZBarSymbol.QRCODE])
            
            if not decoded_objects:
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                enhanced = clahe.apply(gray if len(image.shape) == 2 else cv2.cvtColor(image, cv2.COLOR_BGR2GRAY))
                decoded_objects = decode(enhanced, symbols=[ZBarSymbol.QRCODE])
            
            if decoded_objects:
                return decoded_objects[0].data.decode('utf-8'), None
            return None, '未检测到二维码'
        except Exception as e:
            return None, f'二维码识别错误: {str(e)}'
    
    @staticmethod
    def decode_qrcode_from_pdf(pdf_path):
        """从PDF文件中识别二维码"""
        try:
            doc = fitz.open(pdf_path)
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                
                img_data = pix.tobytes("png")
                nparr = np.frombuffer(img_data, np.uint8)
                open_cv_image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                
                decoded_objects = decode(open_cv_image, symbols=[ZBarSymbol.QRCODE])
                
                if not decoded_objects:
                    gray = cv2.cvtColor(open_cv_image, cv2.COLOR_BGR2GRAY)
                    decoded_objects = decode(gray, symbols=[ZBarSymbol.QRCODE])
                
                if not decoded_objects:
                    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                    enhanced = clahe.apply(gray)
                    decoded_objects = decode(enhanced, symbols=[ZBarSymbol.QRCODE])
                
                if decoded_objects:
                    doc.close()
                    return decoded_objects[0].data.decode('utf-8'), None
            
            doc.close()
            return None, 'PDF中未检测到二维码'
        except Exception as e:
            return None, f'PDF处理错误: {str(e)}'
    
    @staticmethod
    def parse_invoice_qrcode(qrcode_data):
        """解析发票二维码数据，支持多种格式"""
        if not qrcode_data:
            return None, '二维码数据为空'
        
        try:
            if qrcode_data.startswith('http'):
                xml_content, error = InvoiceParser.fetch_invoice_xml_from_url(qrcode_data)
                if xml_content:
                    result = InvoiceParser.parse_xml_to_dict(xml_content)
                    if result:
                        return result, None
                
                return InvoiceParser.parse_url_format(qrcode_data)
            
            if '<' in qrcode_data and '>' in qrcode_data and '<?xml' in qrcode_data.lower():
                return InvoiceParser.parse_xml_to_dict(qrcode_data), None
            
            if re.match(r'^[A-Za-z0-9+/=]+$', qrcode_data) and len(qrcode_data) > 50:
                result = InvoiceParser.decode_base64_to_xml(qrcode_data)
                if result:
                    if '<' in result and '>' in result:
                        return InvoiceParser.parse_xml_to_dict(result), None
            
            parts = qrcode_data.split(',')
            if len(parts) >= 4:
                return InvoiceParser.parse_csv_format(parts), None
            
            return None, f'无法识别的二维码格式: {qrcode_data[:100]}...'
        except Exception as e:
            return None, f'解析错误: {str(e)}'
    
    @staticmethod
    def parse_url_format(url):
        """解析URL格式的二维码"""
        invoice_data = {}
        
        patterns = {
            '发票代码': [r'fpdm=([^&]+)', r'd=([^&]+)', r'code=([^&]+)'],
            '发票号码': [r'fphm=([^&]+)', r'e=([^&]+)', r'no=([^&]+)', r'number=([^&]+)'],
            '校验码': [r'jym=([^&]+)', r'k=([^&]+)', r'checkCode=([^&]+)', r'check=([^&]+)'],
            '开票日期': [r'kprq=([^&]+)', r't=([^&]+)', r'date=([^&]+)', r'time=([^&]+)'],
            '合计金额': [r'kjje=([^&]+)', r'm=([^&]+)', r'amount=([^&]+)', r'money=([^&]+)'],
            '价税合计': [r'kpje=([^&]+)', r'total=([^&]+)'],
            '购买方名称': [r'gmfmc=([^&]+)', r'buyerName=([^&]+)'],
            '销售方名称': [r'xsfmc=([^&]+)', r'sellerName=([^&]+)'],
        }
        
        for field, pats in patterns.items():
            for pat in pats:
                match = re.search(pat, url, re.IGNORECASE)
                if match:
                    invoice_data[field] = match.group(1)
                    break
        
        if '开票日期' in invoice_data:
            date_str = invoice_data['开票日期']
            if len(date_str) == 8 and date_str.isdigit():
                invoice_data['开票日期'] = f'{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}'
            elif len(date_str) == 10 and '-' in date_str:
                pass
        
        if invoice_data:
            invoice_data['发票类型'] = '增值税发票'
            return invoice_data, None
        
        return None, 'URL中未找到发票信息'
    
    @staticmethod
    def parse_csv_format(parts):
        """解析CSV格式的二维码数据
        
        全电发票格式：版本号,发票种类,发票代码,发票号码,金额,开票日期,校验码,加密字段
        例如：01,32,,26427000000316294884,100.76,20260330,,5F98
        
        传统格式：发票类型代码,发票代码,发票号码,金额,日期,校验码(后6位)
        或：发票类型,发票代码,发票号码,金额,税额,日期,校验码
        """
        invoice_data = {}
        
        try:
            parts = [p.strip() for p in parts]
            
            invoice_type_map = {
                '01': '增值税专用发票',
                '02': '货运专票',
                '03': '机动车发票',
                '04': '增值税普通发票',
                '10': '电子普通发票',
                '11': '卷式发票',
                '14': '电子专票',          
                '31': '增值税普通发票（电子）',
                '32': '全电发票（电子发票）',
                '33': '全电发票（增值税专用发票）',
                '51': '铁路电子客票',
            }
            
            if len(parts) >= 6:
                if parts[0] == '01' and parts[1].isdigit():
                    invoice_data['版本号'] = parts[0]
                    type_code = parts[1]
                    invoice_data['发票类型代码'] = type_code
                    invoice_data['发票类型'] = invoice_type_map.get(type_code, f'发票类型{type_code}')
                    invoice_data['发票代码'] = parts[2] if len(parts) > 2 else ''
                    invoice_data['发票号码'] = parts[3] if len(parts) > 3 else ''
                    invoice_data['价税合计'] = parts[4] if len(parts) > 4 else ''
                    invoice_data['开票日期'] = parts[5] if len(parts) > 5 else ''
                    if len(parts) > 6:
                        invoice_data['校验码'] = parts[6]
                    if len(parts) > 7:
                        invoice_data['加密字段'] = parts[7]
                elif parts[0].isdigit() and len(parts[0]) == 2:
                    type_code = parts[0]
                    invoice_data['发票类型代码'] = type_code
                    invoice_data['发票类型'] = invoice_type_map.get(type_code, f'发票类型{type_code}')
                    invoice_data['发票代码'] = parts[1]
                    invoice_data['发票号码'] = parts[2]
                    invoice_data['价税合计'] = parts[3]
                    invoice_data['开票日期'] = parts[4]
                    invoice_data['校验码'] = parts[5] if len(parts) > 5 else ''
                else:
                    invoice_data['发票代码'] = parts[0]
                    invoice_data['发票号码'] = parts[1]
                    invoice_data['价税合计'] = parts[2]
                    invoice_data['开票日期'] = parts[3]
                    if len(parts) > 4:
                        invoice_data['校验码'] = parts[4]
                    if len(parts) > 5:
                        invoice_data['合计税额'] = parts[5]
                    invoice_data['发票类型'] = '增值税发票'
            elif len(parts) >= 4:
                invoice_data['发票代码'] = parts[0]
                invoice_data['发票号码'] = parts[1]
                invoice_data['价税合计'] = parts[2]
                invoice_data['开票日期'] = parts[3]
                invoice_data['发票类型'] = '增值税发票'
            
            if '开票日期' in invoice_data:
                date_str = invoice_data['开票日期']
                if len(date_str) == 8 and date_str.isdigit():
                    invoice_data['开票日期'] = f'{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}'
            
            if '发票类型' not in invoice_data:
                invoice_data['发票类型'] = '增值税发票'
            
            if invoice_data.get('发票类型代码') == '51':
                invoice_data['销售方名称'] = '中国国家铁路集团有限公司'
                if '价税合计' in invoice_data:
                    invoice_data['合计金额'] = invoice_data['价税合计']
                invoice_data['合计税额'] = '0'
            
            return invoice_data
        except Exception as e:
            return {'发票类型': '增值税发票', '错误': str(e)}
    
    @staticmethod
    def decode_base64_to_xml(base64_data):
        """将Base64数据解码为XML"""
        try:
            decoded_bytes = base64.b64decode(base64_data)
            
            if decoded_bytes[:4] == b'PK\x03\x04':
                with zipfile.ZipFile(io.BytesIO(decoded_bytes), 'r') as zf:
                    xml_files = [f for f in zf.namelist() if f.endswith('.xml')]
                    if xml_files:
                        content = zf.read(xml_files[0])
                        try:
                            return content.decode('utf-8')
                        except:
                            return content.decode('gbk')
            
            try:
                return decoded_bytes.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    return decoded_bytes.decode('gbk')
                except:
                    return decoded_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            add_log(f"Base64解码错误: {e}")
            return None
    
    @staticmethod
    def parse_xml_to_dict(xml_content):
        """解析XML内容为字典"""
        if not xml_content:
            return None
        
        try:
            root = ET.fromstring(xml_content)
            
            namespace = ''
            if root.tag.startswith('{'):
                namespace = root.tag.split('}')[0] + '}'
            
            def find_element(parent, *tags):
                for tag in tags:
                    elem = parent.find(namespace + tag) if namespace else parent.find(tag)
                    if elem is not None and elem.text:
                        return elem.text.strip()
                return ''
            
            def find_all_elements(parent, *tags):
                for tag in tags:
                    elems = parent.findall('.//' + namespace + tag) if namespace else parent.findall('.//' + tag)
                    if elems:
                        return elems
                    elems = parent.findall(namespace + tag) if namespace else parent.findall(tag)
                    if elems:
                        return elems
                return []
            
            def get_attr(parent, attr_name):
                return parent.get(attr_name, '') if parent is not None else ''
            
            invoice_data = {}
            
            if root.tag == 'EInvoice' or root.find('EInvoiceData') is not None or root.find('.//SellerInformation') is not None:
                invoice_data['发票类型'] = '电子发票'
                
                seller_info = root.find('.//SellerInformation')
                if seller_info is not None:
                    invoice_data['销售方名称'] = find_element(seller_info, 'SellerName')
                    invoice_data['销售方纳税人识别号'] = find_element(seller_info, 'SellerIdNum')
                    seller_addr = find_element(seller_info, 'SellerAddr')
                    seller_tel = find_element(seller_info, 'SellerTelNum')
                    invoice_data['销售方地址电话'] = f"{seller_addr} {seller_tel}".strip()
                
                buyer_info = root.find('.//BuyerInformation')
                if buyer_info is not None:
                    invoice_data['购买方名称'] = find_element(buyer_info, 'BuyerName')
                    invoice_data['购买方纳税人识别号'] = find_element(buyer_info, 'BuyerIdNum')
                    invoice_data['购买方地址电话'] = find_element(buyer_info, 'BuyerAddr')
                    invoice_data['购买方开户行及账号'] = find_element(buyer_info, 'BuyerBankAccount')
                
                basic_info = root.find('.//BasicInformation')
                if basic_info is not None:
                    invoice_data['合计金额'] = find_element(basic_info, 'TotalAmWithoutTax')
                    invoice_data['合计税额'] = find_element(basic_info, 'TotalTaxAm')
                    invoice_data['价税合计'] = find_element(basic_info, 'TotalTax-includedAmount')
                    invoice_data['价税合计大写'] = find_element(basic_info, 'TotalTax-includedAmountInChinese')
                    invoice_data['开票人'] = find_element(basic_info, 'Drawer')
                    invoice_data['开票日期'] = find_element(basic_info, 'RequestTime')
                
                tax_info = root.find('.//TaxSupervisionInfo')
                if tax_info is not None:
                    invoice_data['发票号码'] = find_element(tax_info, 'InvoiceNumber')
                    invoice_data['开票日期'] = find_element(tax_info, 'IssueTime') or invoice_data.get('开票日期', '')
                    invoice_data['税务机关'] = find_element(tax_info, 'TaxBureauName')
                
                header_info = root.find('.//Header')
                if header_info is not None:
                    invoice_data['电子发票ID'] = find_element(header_info, 'EIid')
                
                goods_list = []
                items = root.findall('.//IssuItemInformation')
                for item in items:
                    goods = {
                        '货物名称': find_element(item, 'ItemName'),
                        '规格型号': find_element(item, 'SpecMod'),
                        '单位': find_element(item, 'MeaUnits'),
                        '数量': find_element(item, 'Quantity'),
                        '单价': find_element(item, 'UnPrice'),
                        '金额': find_element(item, 'Amount'),
                        '税率': find_element(item, 'TaxRate'),
                        '税额': find_element(item, 'ComTaxAm'),
                    }
                    if any(goods.values()):
                        goods_list.append(goods)
                
                invoice_data['商品明细'] = goods_list
                return invoice_data
            
            invoice_data = {
                '发票类型': get_attr(root, '发票类型') or get_attr(root, 'invoiceType') or find_element(root, '发票类型名称', 'invoiceType', 'InvoiceType') or '增值税发票',
                '发票代码': find_element(root, '发票代码', 'invoiceCode', 'fpdm', 'InvoiceCode'),
                '发票号码': find_element(root, '发票号码', 'invoiceNumber', 'fphm', 'InvoiceNumber', 'InvoiceNo'),
                '开票日期': find_element(root, '开票日期', '开票时间', 'issueDate', 'kprq', 'IssueDate', 'IssueTime'),
                '校验码': find_element(root, '校验码', 'checkCode', 'jym', 'CheckCode'),
                '机器编号': find_element(root, '机器编号', 'machineNumber', 'MachineNo'),
                '购买方名称': find_element(root, '购买方名称', 'buyerName', 'gmfmc', 'BuyerName', 'Buyer'),
                '购买方纳税人识别号': find_element(root, '购买方纳税人识别号', 'buyerTaxId', 'gmfsbh', 'BuyerTaxID', 'BuyerCode'),
                '购买方地址电话': find_element(root, '购买方地址电话', 'buyerAddressPhone', 'gmfdzdh', 'BuyerAddressPhone'),
                '购买方开户行及账号': find_element(root, '购买方开户行及账号', 'buyerBankAccount', 'gmfyhzh', 'BuyerBankAccount'),
                '销售方名称': find_element(root, '销售方名称', 'sellerName', 'xsfmc', 'SellerName', 'Seller'),
                '销售方纳税人识别号': find_element(root, '销售方纳税人识别号', 'sellerTaxId', 'xsfsbh', 'SellerTaxID', 'SellerCode'),
                '销售方地址电话': find_element(root, '销售方地址电话', 'sellerAddressPhone', 'xsfdzdh', 'SellerAddressPhone'),
                '销售方开户行及账号': find_element(root, '销售方开户行及账号', 'sellerBankAccount', 'xsfyhzh', 'SellerBankAccount'),
                '合计金额': find_element(root, '合计金额', 'totalAmount', 'hjje', 'TotalAmount', 'Amount'),
                '合计税额': find_element(root, '合计税额', 'totalTax', 'hjse', 'TotalTax', 'Tax'),
                '价税合计': find_element(root, '价税合计', 'totalPriceTax', 'jshj', 'TotalPriceTax', 'Total'),
                '价税合计大写': find_element(root, '价税合计大写', 'totalPriceTaxCN', 'jshjdx', 'TotalPriceTaxCN'),
                '收款人': find_element(root, '收款人', 'payee', 'skr', 'Payee'),
                '复核人': find_element(root, '复核人', 'reviewer', 'fhr', 'Reviewer'),
                '开票人': find_element(root, '开票人', 'issuer', 'kpr', 'Issuer', 'Drawer'),
                '备注': find_element(root, '备注', 'remarks', 'bz', 'Remarks', 'Memo'),
            }
            
            goods_list = []
            
            item_tags = ['货物或应税劳务名称', '商品明细', 'invoiceItem', 'hwxx', 'xmx', 'Item', 'Goods', 'InvoiceItem', 'TaxRateItem']
            for item_tag in item_tags:
                items = find_all_elements(root, item_tag)
                if items:
                    for item in items:
                        goods = {
                            '货物名称': find_element(item, '货物或应税劳务名称', '商品名称', 'goodsName', 'hwmc', 'xmmc', 'GoodsName', 'ItemName', 'Name'),
                            '规格型号': find_element(item, '规格型号', 'specification', 'ggxh', 'Specification', 'Model'),
                            '单位': find_element(item, '单位', 'unit', 'dw', 'Unit', 'MeasureUnit'),
                            '数量': find_element(item, '数量', 'quantity', 'sl', 'Quantity', 'Num'),
                            '单价': find_element(item, '单价', 'unitPrice', 'dj', 'UnitPrice', 'Price'),
                            '金额': find_element(item, '金额', 'amount', 'je', 'Amount', 'Money'),
                            '税率': find_element(item, '税率', 'taxRate', 'slv', 'TaxRate'),
                            '税额': find_element(item, '税额', 'tax', 'se', 'Tax', 'TaxAmount'),
                        }
                        if any(goods.values()):
                            goods_list.append(goods)
                    if goods_list:
                        break
            
            invoice_data['商品明细'] = goods_list
            
            return invoice_data
        except Exception as e:
            add_log(f"XML解析错误: {e}")
            return None


class ProcessThread(QThread):
    """处理发票的线程"""
    progress = pyqtSignal(int, int)
    result_ready = pyqtSignal(dict)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, files, max_workers=4):
        super().__init__()
        self.files = files
        self.max_workers = max_workers
        self.results_lock = threading.Lock()
        self.progress_lock = threading.Lock()
        self.completed_count = 0
    
    def process_single_file(self, file_path):
        """处理单个发票文件"""
        add_log(f"\n========== 开始处理: {os.path.basename(file_path)} ==========", file_path)
        
        qrcode_data, qrcode_error = InvoiceParser.decode_qrcode(file_path)
        if not qrcode_data:
            add_log(f"二维码识别失败: {qrcode_error or '无法识别二维码'}", file_path)
            result = {
                '文件名': os.path.basename(file_path),
                '文件路径': file_path,
                '状态': '失败',
                '识别日志': get_file_logs(file_path)
            }
            return result
        
        add_log(f"二维码识别成功", file_path)
        invoice_data, parse_error = InvoiceParser.parse_invoice_qrcode(qrcode_data)
        
        if isinstance(invoice_data, str):
            xml_content = invoice_data
            invoice_data = InvoiceParser.parse_xml_to_dict(xml_content)
            if not invoice_data:
                add_log(f"XML解析失败", file_path)
                result = {
                    '文件名': os.path.basename(file_path),
                    '文件路径': file_path,
                    '状态': '失败',
                    '识别日志': get_file_logs(file_path)
                }
                return result
        
        if not invoice_data:
            add_log(f"发票数据解析失败: {parse_error or '无法解析发票数据'}", file_path)
            result = {
                '文件名': os.path.basename(file_path),
                '文件路径': file_path,
                '状态': '失败',
                '识别日志': get_file_logs(file_path)
            }
            return result
        
        if not invoice_data.get('销售方名称') or not invoice_data.get('销售方纳税人识别号') or not invoice_data.get('购买方名称') or not invoice_data.get('购买方纳税人识别号') or not invoice_data.get('合计金额') or not invoice_data.get('合计税额'):
            add_log(f"缺少必要信息，启动OCR补充识别", file_path)
            existing_total_price = invoice_data.get('价税合计')
            ocr_info, ocr_error = InvoiceParser.ocr_extract_seller_info(file_path, existing_total_price)
            if ocr_info:
                if not invoice_data.get('销售方名称') and ocr_info.get('销售方名称'):
                    invoice_data['销售方名称'] = ocr_info['销售方名称']
                if not invoice_data.get('销售方纳税人识别号') and ocr_info.get('销售方纳税人识别号'):
                    invoice_data['销售方纳税人识别号'] = ocr_info['销售方纳税人识别号']
                if not invoice_data.get('购买方名称') and ocr_info.get('购买方名称'):
                    invoice_data['购买方名称'] = ocr_info['购买方名称']
                if not invoice_data.get('购买方纳税人识别号') and ocr_info.get('购买方纳税人识别号'):
                    invoice_data['购买方纳税人识别号'] = ocr_info['购买方纳税人识别号']
                if not invoice_data.get('合计金额') and ocr_info.get('合计金额'):
                    invoice_data['合计金额'] = ocr_info['合计金额']
                if not invoice_data.get('合计税额') and ocr_info.get('合计税额'):
                    invoice_data['合计税额'] = ocr_info['合计税额']
                if not invoice_data.get('价税合计') and ocr_info.get('价税合计'):
                    invoice_data['价税合计'] = ocr_info['价税合计']
        
        if invoice_data.get('购买方名称') == '深圳市城图科技有限公司' and invoice_data.get('购买方纳税人识别号') != '91440300665885384A':
            add_log(f"购买方为深圳市城图科技有限公司，修正纳税人识别号", file_path)
            if invoice_data.get('购买方纳税人识别号'):
                invoice_data['销售方纳税人识别号'] = invoice_data['购买方纳税人识别号']
            invoice_data['购买方纳税人识别号'] = '91440300665885384A'
        
        invoice_data['文件名'] = os.path.basename(file_path)
        invoice_data['文件路径'] = file_path
        invoice_data['状态'] = '成功'
        add_log(f"识别成功", file_path)
        invoice_data['识别日志'] = get_file_logs(file_path)
        
        return invoice_data
    
    def run(self):
        """使用多线程并发处理发票"""
        results = []
        total = len(self.files)
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_file = {executor.submit(self.process_single_file, file_path): file_path 
                            for file_path in self.files}
            
            for future in as_completed(future_to_file):
                try:
                    result = future.result()
                    
                    with self.results_lock:
                        results.append(result)
                    
                    self.result_ready.emit(result)
                    
                    with self.progress_lock:
                        self.completed_count += 1
                        self.progress.emit(self.completed_count, total)
                
                except Exception as e:
                    file_path = future_to_file[future]
                    add_log(f"处理文件出错: {str(e)}", file_path)
                    error_result = {
                        '文件名': os.path.basename(file_path),
                        '文件路径': file_path,
                        '状态': '失败',
                        '识别日志': get_file_logs(file_path)
                    }
                    with self.results_lock:
                        results.append(error_result)
                    self.result_ready.emit(error_result)
        
        self.finished.emit(results)


class InvoiceMainWindow(QMainWindow):
    """发票识别主窗口"""
    
    def __init__(self):
        super().__init__()
        self.file_list = []
        self.invoice_results = []
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle('发票识别工具')
        self.setGeometry(100, 100, 1400, 900)
        
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.png')
        if os.path.exists(logo_path):
            self.setWindowIcon(QIcon(logo_path))
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        
        toolbar_widget = QWidget()
        toolbar_layout = QHBoxLayout(toolbar_widget)
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        
        self.btn_add = QPushButton('添加文件')
        self.btn_add.clicked.connect(self.add_files)
        self.btn_add.setMinimumHeight(40)
        self.btn_add.setMinimumWidth(100)
        
        self.btn_add_folder = QPushButton('添加文件夹')
        self.btn_add_folder.clicked.connect(self.add_folder)
        self.btn_add_folder.setMinimumHeight(40)
        self.btn_add_folder.setMinimumWidth(100)
        
        self.btn_remove = QPushButton('删除选中')
        self.btn_remove.clicked.connect(self.remove_selected)
        self.btn_remove.setMinimumHeight(40)
        self.btn_remove.setMinimumWidth(100)
        
        self.btn_clear = QPushButton('清空列表')
        self.btn_clear.clicked.connect(self.clear_list)
        self.btn_clear.setMinimumHeight(40)
        self.btn_clear.setMinimumWidth(100)
        
        self.btn_process = QPushButton('开始识别')
        self.btn_process.clicked.connect(self.process_files)
        self.btn_process.setMinimumHeight(40)
        self.btn_process.setMinimumWidth(120)
        self.btn_process.setStyleSheet('background-color: #4CAF50; color: white; font-size: 14px; font-weight: bold;')
        
        self.btn_export = QPushButton('导出Excel')
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_export.setMinimumHeight(40)
        self.btn_export.setMinimumWidth(100)
        self.btn_export.setEnabled(False)
        
        self.btn_view_log = QPushButton('查看日志')
        self.btn_view_log.clicked.connect(self.view_log)
        self.btn_view_log.setMinimumHeight(40)
        self.btn_view_log.setMinimumWidth(80)
        
        self.btn_clear_results = QPushButton('清空结果')
        self.btn_clear_results.clicked.connect(self.clear_results)
        self.btn_clear_results.setMinimumHeight(40)
        self.btn_clear_results.setMinimumWidth(100)
        self.btn_clear_results.setEnabled(False)
        
        self.btn_retry_failed = QPushButton('重试失败')
        self.btn_retry_failed.clicked.connect(self.retry_failed)
        self.btn_retry_failed.setMinimumHeight(40)
        self.btn_retry_failed.setMinimumWidth(100)
        self.btn_retry_failed.setEnabled(False)
        
        toolbar_layout.addWidget(self.btn_add)
        toolbar_layout.addWidget(self.btn_add_folder)
        toolbar_layout.addWidget(self.btn_remove)
        toolbar_layout.addWidget(self.btn_clear)
        toolbar_layout.addSpacing(20)
        toolbar_layout.addWidget(self.btn_process)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.btn_export)
        toolbar_layout.addWidget(self.btn_view_log)
        toolbar_layout.addWidget(self.btn_retry_failed)
        toolbar_layout.addWidget(self.btn_clear_results)
        
        main_layout.addWidget(toolbar_widget)
        
        file_section = QWidget()
        file_layout = QVBoxLayout(file_section)
        file_layout.setContentsMargins(0, 0, 0, 0)
        file_layout.setSpacing(5)
        
        file_header = QHBoxLayout()
        file_header.addWidget(QLabel('待识别文件列表:'))
        self.file_count_label = QLabel('共 0 个文件')
        file_header.addWidget(self.file_count_label)
        file_header.addStretch()
        file_layout.addLayout(file_header)
        
        self.file_listbox = QListWidget()
        self.file_listbox.setSelectionMode(QListWidget.ExtendedSelection)
        self.file_listbox.setAlternatingRowColors(True)
        self.file_listbox.setMinimumHeight(150)
        file_layout.addWidget(self.file_listbox)
        
        main_layout.addWidget(file_section)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        result_section = QWidget()
        result_layout = QVBoxLayout(result_section)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(5)
        
        result_header = QHBoxLayout()
        result_header.addWidget(QLabel('识别结果:'))
        self.result_count_label = QLabel('共 0 条记录')
        result_header.addWidget(self.result_count_label)
        result_header.addStretch()
        result_layout.addLayout(result_header)
        
        self.result_table = QTableWidget()
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed)
        self.result_table.cellClicked.connect(self.on_cell_clicked)
        self.result_table.cellChanged.connect(self.on_cell_changed)
        result_layout.addWidget(self.result_table)
        
        main_layout.addWidget(result_section, 1)
        
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage('就绪')
    
    def on_cell_clicked(self, row, col):
        """单击单元格"""
        headers = [
            '文件名', '状态', '发票号码', '开票日期',
            '购买方名称', '购买方纳税人识别号', '销售方名称', '销售方纳税人识别号',
            '合计金额', '合计税额', '价税合计', '识别日志'
        ]
        
        if row < len(self.invoice_results):
            result = self.invoice_results[row]
            header = headers[col] if col < len(headers) else ''
            
            if header == '文件名':
                file_path = result.get('文件路径', '')
                if file_path and os.path.exists(file_path):
                    import subprocess
                    import platform
                    try:
                        if platform.system() == 'Windows':
                            os.startfile(file_path)
                        elif platform.system() == 'Darwin':
                            subprocess.run(['open', file_path])
                        else:
                            subprocess.run(['xdg-open', file_path])
                    except Exception as e:
                        QMessageBox.warning(self, '提示', f'无法打开文件: {str(e)}')
                else:
                    QMessageBox.warning(self, '提示', '文件路径不存在')
            elif header == '识别日志':
                item = self.result_table.item(row, col)
                log_content = item.data(Qt.UserRole) if item else '无日志信息'
                self.show_invoice_log(log_content, result.get('文件名', ''))
    
    def on_cell_changed(self, row, col):
        """单元格内容改变时保存"""
        if row < len(self.invoice_results):
            headers = [
                '文件名', '状态', '发票号码', '开票日期',
                '购买方名称', '购买方纳税人识别号', '销售方名称', '销售方纳税人识别号',
                '合计金额', '合计税额', '价税合计', '识别日志'
            ]
            if col < len(headers):
                item = self.result_table.item(row, col)
                if item:
                    self.invoice_results[row][headers[col]] = item.text()
    
    def view_log(self):
        """查看识别日志"""
        dialog = QDialog(self)
        dialog.setWindowTitle('识别日志')
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(get_logs())
        text_edit.setStyleSheet('font-family: Consolas, monospace; font-size: 12px;')
        layout.addWidget(text_edit)
        
        btn_layout = QHBoxLayout()
        btn_clear = QPushButton('清空日志')
        btn_clear.clicked.connect(lambda: (clear_logs(), text_edit.setPlainText('')))
        btn_close = QPushButton('关闭')
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_clear)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def show_invoice_log(self, log_content, file_name=''):
        """显示单个发票的识别日志"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f'识别日志 - {file_name}')
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(log_content if log_content else '无日志信息')
        text_edit.setStyleSheet('font-family: Consolas, monospace; font-size: 12px;')
        layout.addWidget(text_edit)
        
        btn_layout = QHBoxLayout()
        btn_close = QPushButton('关闭')
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def add_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            '选择发票文件',
            '',
            '图片文件 (*.png *.jpg *.jpeg *.bmp *.pdf);;所有文件 (*.*)'
        )
        if files:
            self.add_files_to_list(files)
    
    def add_folder(self):
        folder = QFileDialog.getExistingDirectory(self, '选择文件夹')
        if folder:
            files = []
            for ext in ['*.png', '*.jpg', '*.jpeg', '*.bmp', '*.pdf']:
                files.extend(self.get_files_by_pattern(folder, ext))
            if files:
                self.add_files_to_list(files)
    
    def get_files_by_pattern(self, folder, pattern):
        import glob
        return glob.glob(os.path.join(folder, pattern))
    
    def add_files_to_list(self, files):
        for f in files:
            if f not in self.file_list:
                self.file_list.append(f)
                self.file_listbox.addItem(f)
        self.update_file_count()
    
    def remove_selected(self):
        selected_items = self.file_listbox.selectedItems()
        if not selected_items:
            return
        
        reply = QMessageBox.question(
            self, '确认删除',
            f'确定要删除选中的 {len(selected_items)} 个文件吗？',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        
        for item in selected_items:
            row = self.file_listbox.row(item)
            self.file_listbox.takeItem(row)
            if item.text() in self.file_list:
                self.file_list.remove(item.text())
        self.update_file_count()
    
    def clear_list(self):
        if not self.file_list:
            return
        
        reply = QMessageBox.question(
            self, '确认清空',
            f'确定要清空文件列表吗？\n当前共有 {len(self.file_list)} 个文件。',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        
        self.file_list.clear()
        self.file_listbox.clear()
        self.update_file_count()
    
    def clear_results(self):
        if not self.invoice_results:
            return
        
        reply = QMessageBox.question(
            self, '确认清空',
            f'确定要清空识别结果吗？\n当前共有 {len(self.invoice_results)} 条记录。',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        
        self.invoice_results.clear()
        self.result_table.setRowCount(0)
        self.result_count_label.setText('共 0 条记录')
        self.btn_export.setEnabled(False)
        self.btn_clear_results.setEnabled(False)
        self.btn_retry_failed.setEnabled(False)
        self.status_bar.showMessage('识别结果已清空')
    
    def retry_failed(self):
        """重新识别失败的文件"""
        failed_files = [r.get('文件路径') for r in self.invoice_results if r.get('状态') == '失败']
        
        if not failed_files:
            QMessageBox.information(self, '提示', '没有失败的记录需要重新识别！')
            return
        
        reply = QMessageBox.question(
            self, '确认重试',
            f'确定要重新识别 {len(failed_files)} 个失败的文件吗？',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.No:
            return
        
        failed_indices = [i for i, r in enumerate(self.invoice_results) if r.get('状态') == '失败']
        for i in reversed(failed_indices):
            if i < self.result_table.rowCount():
                self.result_table.removeRow(i)
        
        self.invoice_results = [r for r in self.invoice_results if r.get('状态') == '成功']
        
        if not self.invoice_results:
            self.result_count_label.setText('共 0 条记录')
            self.btn_export.setEnabled(False)
            self.btn_clear_results.setEnabled(False)
        
        self.status_bar.showMessage(f'正在重新识别 {len(failed_files)} 个失败的文件...')
        
        self.btn_process.setEnabled(False)
        self.btn_add.setEnabled(False)
        self.btn_add_folder.setEnabled(False)
        self.btn_remove.setEnabled(False)
        self.btn_clear.setEnabled(False)
        self.btn_retry_failed.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        self.process_thread = ProcessThread(failed_files)
        self.process_thread.progress.connect(self.update_progress)
        self.process_thread.result_ready.connect(self.add_result)
        self.process_thread.finished.connect(self.on_process_finished)
        self.process_thread.start()
    
    def update_file_count(self):
        self.file_count_label.setText(f'共 {len(self.file_list)} 个文件')
    
    def process_files(self):
        if not self.file_list:
            QMessageBox.warning(self, '提示', '请先添加发票文件！')
            return
        
        existing_files = {r.get('文件路径') for r in self.invoice_results if r.get('文件路径')}
        new_files = [f for f in self.file_list if f not in existing_files]
        
        if not new_files:
            QMessageBox.information(self, '提示', '所有文件都已识别过！')
            return
        
        if len(new_files) < len(self.file_list):
            skip_count = len(self.file_list) - len(new_files)
            reply = QMessageBox.question(
                self, '提示',
                f'有 {skip_count} 个文件已识别过，将跳过这些文件。\n是否继续识别 {len(new_files)} 个新文件？',
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
        
        self.btn_process.setEnabled(False)
        self.btn_add.setEnabled(False)
        self.btn_add_folder.setEnabled(False)
        self.btn_remove.setEnabled(False)
        self.btn_clear.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_bar.showMessage('正在识别...')
        
        self.process_thread = ProcessThread(new_files)
        self.process_thread.progress.connect(self.update_progress)
        self.process_thread.result_ready.connect(self.add_result)
        self.process_thread.finished.connect(self.on_process_finished)
        self.process_thread.start()
    
    def update_progress(self, current, total):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_bar.showMessage(f'正在识别: {current}/{total}')
    
    def add_result(self, result):
        """实时添加一条识别结果"""
        self.invoice_results.append(result)
        
        headers = [
            '文件名', '状态', '发票号码', '开票日期',
            '购买方名称', '购买方纳税人识别号', '销售方名称', '销售方纳税人识别号',
            '合计金额', '合计税额', '价税合计', '识别日志'
        ]
        
        if self.result_table.columnCount() == 0:
            self.result_table.setColumnCount(len(headers))
            self.result_table.setHorizontalHeaderLabels(headers)
            status_col_index = headers.index('状态')
            self.result_table.setItemDelegateForColumn(status_col_index, StatusTagDelegate())
        
        self.result_table.blockSignals(True)
        
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        for col, header in enumerate(headers):
            value = result.get(header, '')
            item = QTableWidgetItem(str(value) if value else '')
            if header == '状态':
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            elif header == '文件名':
                item.setForeground(QColor(0, 0, 255))
                font = item.font()
                font.setUnderline(True)
                item.setFont(font)
                file_path = result.get('文件路径', '')
                if file_path:
                    item.setToolTip(f'双击打开: {file_path}')
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            elif header == '识别日志':
                item.setData(Qt.UserRole, value)
                item.setText('查看日志')
                item.setForeground(QColor(0, 0, 255))
                font = item.font()
                font.setUnderline(True)
                item.setFont(font)
                item.setToolTip('双击查看详细识别日志')
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.result_table.setItem(row, col, item)
        
        self.result_table.blockSignals(False)
        
        header = self.result_table.horizontalHeader()
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.result_table.resizeColumnsToContents()
        
        self.result_count_label.setText(f'共 {len(self.invoice_results)} 条记录')
        
        success_count = sum(1 for r in self.invoice_results if r.get('状态') == '成功')
        failed_count = len(self.invoice_results) - success_count
        self.status_bar.showMessage(f'识别中: 成功 {success_count} 个, 失败 {failed_count} 个')
        
        if self.invoice_results:
            self.btn_export.setEnabled(True)
            self.btn_clear_results.setEnabled(True)
            if failed_count > 0:
                self.btn_retry_failed.setEnabled(True)
    
    def on_process_finished(self, results):
        for result in results:
            if result not in self.invoice_results:
                self.invoice_results.append(result)
        
        self.btn_process.setEnabled(True)
        self.btn_add.setEnabled(True)
        self.btn_add_folder.setEnabled(True)
        self.btn_remove.setEnabled(True)
        self.btn_clear.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.btn_export.setEnabled(True)
        self.btn_clear_results.setEnabled(True)
        
        success_count = sum(1 for r in self.invoice_results if r.get('状态') == '成功')
        failed_count = len(self.invoice_results) - success_count
        self.status_bar.showMessage(f'识别完成: 成功 {success_count} 个, 失败 {failed_count} 个')
        
        if failed_count > 0:
            self.btn_retry_failed.setEnabled(True)
    
    def display_results(self, results):
        if not results:
            return
        
        headers = [
            '文件名', '状态', '发票号码', '开票日期',
            '购买方名称', '购买方纳税人识别号', '销售方名称', '销售方纳税人识别号',
            '合计金额', '合计税额', '价税合计', '识别日志'
        ]
        
        self.result_table.setColumnCount(len(headers))
        self.result_table.setHorizontalHeaderLabels(headers)
        self.result_table.setRowCount(len(results))
        
        for row, result in enumerate(results):
            for col, header in enumerate(headers):
                value = result.get(header, '')
                item = QTableWidgetItem(str(value) if value else '')
                if header == '状态':
                    if value == '成功':
                        item.setBackground(Qt.green)
                    else:
                        item.setBackground(Qt.red)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                elif header == '文件名':
                    item.setForeground(QColor(0, 0, 255))
                    font = item.font()
                    font.setUnderline(True)
                    item.setFont(font)
                    file_path = result.get('文件路径', '')
                    if file_path:
                        item.setToolTip(f'双击打开: {file_path}')
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                elif header == '识别日志':
                    item.setData(Qt.UserRole, value)
                    item.setText('查看日志')
                    item.setForeground(QColor(0, 0, 255))
                    font = item.font()
                    font.setUnderline(True)
                    item.setFont(font)
                    item.setToolTip('双击查看详细识别日志')
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.result_table.setItem(row, col, item)
        
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.result_table.horizontalHeader().setStretchLastSection(True)
        
        self.result_count_label.setText(f'共 {len(results)} 条记录')
    
    def export_to_excel(self):
        if not self.invoice_results:
            QMessageBox.warning(self, '提示', '没有可导出的数据！')
            return
        
        all_headers = [
            '文件名', '状态', '发票号码', '开票日期',
            '购买方名称', '购买方纳税人识别号', '销售方名称', '销售方纳税人识别号',
            '合计金额', '合计税额', '价税合计', '识别日志'
        ]
        
        default_selected = all_headers.copy()
        default_selected.remove('状态')
        default_selected.remove('识别日志')
        
        dialog = FieldSelectDialog(all_headers, default_selected, self)
        if dialog.exec_() != QDialog.Accepted:
            return
        
        selected_headers = dialog.get_selected_fields()
        if not selected_headers:
            QMessageBox.warning(self, '提示', '请至少选择一个字段！')
            return
        
        default_name = f'发票识别结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            '保存Excel文件',
            default_name,
            'Excel文件 (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '发票信息'
            
            header_font = ExcelFont(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(selected_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row, result in enumerate(self.invoice_results, 2):
                for col, header in enumerate(selected_headers, 1):
                    value = result.get(header, '')
                    cell = ws.cell(row=row, column=col, value=str(value) if value else '')
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            ws.freeze_panes = 'A2'
            
            if any(r.get('商品明细') for r in self.invoice_results if r.get('状态') == '成功'):
                ws_detail = wb.create_sheet(title='商品明细')
                
                detail_headers = [
                    '发票号码', '货物名称', '规格型号', '单位', '数量', '单价', '金额', '税率', '税额'
                ]
                
                for col, header in enumerate(detail_headers, 1):
                    cell = ws_detail.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                detail_row = 2
                for result in self.invoice_results:
                    if result.get('状态') == '成功' and result.get('商品明细'):
                        invoice_no = result.get('发票号码', '')
                        for goods in result['商品明细']:
                            ws_detail.cell(row=detail_row, column=1, value=invoice_no).border = thin_border
                            for col, key in enumerate(['货物名称', '规格型号', '单位', '数量', '单价', '金额', '税率', '税额'], 2):
                                cell = ws_detail.cell(row=detail_row, column=col, value=str(goods.get(key, '')))
                                cell.border = thin_border
                            detail_row += 1
                
                for col in ws_detail.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_detail.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, '成功', f'导出成功！\n文件保存在: {file_path}')
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出失败: {str(e)}')


class FieldSelectDialog(QDialog):
    """字段选择对话框"""
    
    def __init__(self, all_fields, default_selected, parent=None):
        super().__init__(parent)
        self.setWindowTitle('选择导出字段')
        self.setMinimumWidth(400)
        self.setMinimumHeight(500)
        self.all_fields = all_fields
        
        layout = QVBoxLayout(self)
        
        label = QLabel('请选择要导出的字段：')
        layout.addWidget(label)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        self.checkboxes = {}
        for field in all_fields:
            cb = QCheckBox(field)
            cb.setChecked(field in default_selected)
            self.checkboxes[field] = cb
            scroll_layout.addWidget(cb)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        
        btn_select_all = QPushButton('全选')
        btn_select_all.clicked.connect(self.select_all)
        btn_layout.addWidget(btn_select_all)
        
        btn_deselect_all = QPushButton('取消全选')
        btn_deselect_all.clicked.connect(self.deselect_all)
        btn_layout.addWidget(btn_deselect_all)
        
        layout.addLayout(btn_layout)
        
        btn_box = QHBoxLayout()
        btn_ok = QPushButton('确定')
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton('取消')
        btn_cancel.clicked.connect(self.reject)
        btn_box.addStretch()
        btn_box.addWidget(btn_ok)
        btn_box.addWidget(btn_cancel)
        layout.addLayout(btn_box)
    
    def select_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(True)
    
    def deselect_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(False)
    
    def get_selected_fields(self):
        return [field for field in self.all_fields if self.checkboxes[field].isChecked()]


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.png')
    if os.path.exists(logo_path):
        app.setWindowIcon(QIcon(logo_path))
    
    font = QFont('Microsoft YaHei', 9)
    app.setFont(font)
    
    window = InvoiceMainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
